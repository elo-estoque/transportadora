import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ==============================================================================
# 1. CONFIGURAÇÕES DE CORES (HEXADECIMAL SEM O #)
# ==============================================================================
# Nota: Para o OpenPyXL, é melhor passar a cor sem o '#'
REGRAS_DE_CORES = {
    'nome_fantasia': 'ADD8E6',       # Azul Claro
    'website': 'E6E6FA',             # Lavanda
    'cnpj': 'FFFFE0',                # Amarelo Claro
    'cidade': 'F0FFF0',              # Verde bem claro
    'estado': 'F5F5DC',              # Bege
    'telefone_1': 'FFB6C1',          # Vermelho Claro
    'email_1': 'FFD700',             # Dourado
    'telefone_2': 'FFC0CB',          # Rosa
    'representante_nome': '98FB98',  # Verde Pálido
    'representante_email': 'FFA07A', # Salmão Claro
    'categoria_nome': 'D3D3D3'       # Cinza Claro
}

ARQUIVO_SAIDA = "Consolidado_Final_Cores_Reais.xlsx"

# ==============================================================================
# 2. CONSOLIDAÇÃO DOS DADOS
# ==============================================================================
print("🔄 Lendo arquivos...")
arquivos = glob.glob("Clientes_*.xlsx")
lista_dfs = []

if not arquivos:
    print("❌ Nenhum arquivo encontrado.")
else:
    for arquivo in arquivos:
        try:
            df = pd.read_excel(arquivo)
            # Cria coluna do vendedor
            nome = os.path.basename(arquivo).replace('.xlsx', '').replace('Clientes_', '')
            df['Vendedor_Origem'] = nome
            lista_dfs.append(df)
        except Exception as e:
            print(f"Erro em {arquivo}: {e}")

    if lista_dfs:
        df_final = pd.concat(lista_dfs, ignore_index=True)

        # ==============================================================================
        # 3. SALVAR PRIMEIRO (SEM CORES AINDA)
        # ==============================================================================
        # Calculamos quantas linhas a legenda vai ocupar para pular no Excel
        linhas_pulo = len(REGRAS_DE_CORES) + 2
        
        print("💾 Salvando estrutura base...")
        # Salvamos o arquivo bruto começando na linha correta
        df_final.to_excel(ARQUIVO_SAIDA, index=False, startrow=linhas_pulo)

        # ==============================================================================
        # 4. PINTURA "FORÇA BRUTA" (OPENPYXL)
        # ==============================================================================
        print("🎨 Pintando células (Modo Compatibilidade Google Sheets)...")
        
        wb = load_workbook(ARQUIVO_SAIDA)
        ws = wb.active

        # 1. Mapear onde está cada coluna (ex: 'telefone_1' é coluna 15?)
        # A linha de cabeçalho é: linhas_pulo + 1 (porque o Excel começa no 1)
        linha_cabecalho = linhas_pulo + 1
        
        mapa_colunas = {}
        for cell in ws[linha_cabecalho]:
            if cell.value:
                mapa_colunas[cell.value] = cell.column

        # 2. Iterar linha por linha e pintar se estiver vazio
        # Começa na linha logo após o cabeçalho
        for row in ws.iter_rows(min_row=linha_cabecalho + 1):
            for col_nome, cor_hex in REGRAS_DE_CORES.items():
                
                # Se a coluna existe no arquivo
                if col_nome in mapa_colunas:
                    col_idx = mapa_colunas[col_nome]
                    cell = row[col_idx - 1] # Pega a célula correta
                    
                    # Verifica se está vazio (None, string vazia ou espaço)
                    valor = cell.value
                    if valor is None or str(valor).strip() == "":
                        # APLICA A COR DE FUNDO (FILL)
                        cell.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

        # ==============================================================================
        # 5. CRIAR A LEGENDA MANUALMENTE
        # ==============================================================================
        ws['A1'] = "LEGENDA DE DADOS FALTANTES:"
        ws['A1'].font = Font(bold=True, size=12)

        linha_atual = 2
        for col_nome, cor_hex in REGRAS_DE_CORES.items():
            cell_texto = ws.cell(row=linha_atual, column=1)
            cell_texto.value = f"Campo '{col_nome}' Vazio"
            
            # Pinta a célula da legenda para referência visual
            cell_cor = ws.cell(row=linha_atual, column=2) # Coluna B pinta a cor
            cell_cor.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
            
            linha_atual += 1

        print("💾 Salvando arquivo final...")
        wb.save(ARQUIVO_SAIDA)
        print("🚀 CONCLUÍDO! O arquivo 'Consolidado_Final_Cores_Reais.xlsx' está pronto.")

    else:
        print("❌ Nada consolidado.")
